#!/usr/bin/env python3
"""
Layer 1.B API experiment runner.

Reads the 02_trial_execution sheet from the populated workbook, executes each trial as an
independent single-prompt API request, and writes the response back to response_text.

No conversation history is used. Each request contains only the current row's prompt_text.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional

from openai import OpenAI
from openpyxl import load_workbook

SHEET_NAME = "02_trial_execution"
REQUIRED_COLUMNS = [
    "trial_id",
    "prompt_id",
    "run_index",
    "random_order",
    "execution_date",
    "model_id",
    "temperature",
    "top_p",
    "system_prompt_used",
    "conversation_history_used",
    "tools_used",
    "browsing_used",
    "memory_used",
    "prompt_text",
    "response_text",
    "response_length_chars",
    "raw_output_saved",
    "notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run Layer 1.B formal API trials from workbook.")
    parser.add_argument("--input", required=True, help="Input workbook path (.xlsx).")
    parser.add_argument("--output", required=True, help="Output workbook path (.xlsx). Can equal --input for in-place update.")
    parser.add_argument("--model", required=True, help="Exact model ID used for execution. Record this in the workbook.")
    parser.add_argument("--max-output-tokens", type=int, default=900, help="Maximum output tokens per response.")
    parser.add_argument("--temperature", type=float, default=0.3, help="Temperature if sampling params are included.")
    parser.add_argument("--top-p", type=float, default=1.0, help="top_p if sampling params are included.")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--omit-sampling-params", action="store_true", help="Do not send temperature/top_p; record model-compatible default.")
    group.add_argument("--include-sampling-params", action="store_true", help="Send temperature/top_p. If unsupported, retry without them.")
    parser.add_argument("--start-order", type=int, default=None, help="Only run trials with random_order >= this value.")
    parser.add_argument("--end-order", type=int, default=None, help="Only run trials with random_order <= this value.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of not-yet-filled trials to process.")
    parser.add_argument("--sleep-seconds", type=float, default=0.5, help="Sleep between successful API calls.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing response_text values.")
    parser.add_argument("--dry-run", action="store_true", help="Print selected trials without calling API or writing output.")
    parser.add_argument("--save-raw-json", action="store_true", help="Save raw API response JSON per trial.")
    parser.add_argument("--raw-json-dir", default="raw_api_json", help="Directory for raw JSON files if --save-raw-json is used.")
    return parser.parse_args()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def get_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = normalize_header(ws.cell(row=1, column=col).value)
        if name:
            headers[name] = col
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns in {SHEET_NAME}: {missing}")
    return headers


def cell(ws, row: int, headers: Dict[str, int], col_name: str):
    return ws.cell(row=row, column=headers[col_name])


def safe_filename(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("_")
    return cleaned[:180] or "trial"


def extract_text(response: Any) -> str:
    # Newer OpenAI Python SDKs expose output_text on Responses API results.
    text = getattr(response, "output_text", None)
    if isinstance(text, str) and text.strip():
        return text

    # Fallback: recursively collect text-like fields from response.model_dump().
    data = response.model_dump() if hasattr(response, "model_dump") else response
    texts = []

    def walk(obj: Any) -> None:
        if isinstance(obj, dict):
            # Prefer obvious text fields.
            if isinstance(obj.get("text"), str):
                texts.append(obj["text"])
            elif isinstance(obj.get("content"), str):
                texts.append(obj["content"])
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(data)
    joined = "\n".join(t for t in texts if t.strip()).strip()
    return joined


def make_response(client: OpenAI, model: str, prompt: str, max_output_tokens: int, include_sampling: bool, temperature: float, top_p: float):
    kwargs: Dict[str, Any] = {
        "model": model,
        "input": prompt,
        "max_output_tokens": max_output_tokens,
    }
    if include_sampling:
        kwargs["temperature"] = temperature
        kwargs["top_p"] = top_p
    return client.responses.create(**kwargs)


def main() -> int:
    args = parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        print(f"ERROR: Input workbook not found: {input_path}", file=sys.stderr)
        return 2

    if not args.dry_run and not os.environ.get("OPENAI_API_KEY"):
        print("ERROR: OPENAI_API_KEY is not set.", file=sys.stderr)
        return 2

    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet not found: {SHEET_NAME}. Available sheets: {wb.sheetnames}", file=sys.stderr)
        return 2
    ws = wb[SHEET_NAME]
    headers = get_headers(ws)

    rows = []
    for row in range(2, ws.max_row + 1):
        prompt_text = cell(ws, row, headers, "prompt_text").value
        if not prompt_text or not str(prompt_text).strip():
            continue
        random_order = cell(ws, row, headers, "random_order").value
        try:
            order_num = int(random_order)
        except Exception:
            continue
        if args.start_order is not None and order_num < args.start_order:
            continue
        if args.end_order is not None and order_num > args.end_order:
            continue
        response_existing = cell(ws, row, headers, "response_text").value
        if response_existing and str(response_existing).strip() and not args.overwrite:
            continue
        rows.append((order_num, row))

    rows.sort(key=lambda x: x[0])
    if args.limit is not None:
        rows = rows[: args.limit]

    print(f"Selected {len(rows)} trials to process.")
    if args.dry_run:
        for order_num, row in rows[:20]:
            trial_id = cell(ws, row, headers, "trial_id").value
            prompt_preview = str(cell(ws, row, headers, "prompt_text").value).replace("\n", " ")[:160]
            print(f"random_order={order_num} row={row} trial_id={trial_id} prompt={prompt_preview!r}")
        print("Dry run complete. No API calls made and no workbook written.")
        return 0

    client = OpenAI()
    raw_dir = Path(args.raw_json_dir)
    if args.save_raw_json:
        raw_dir.mkdir(parents=True, exist_ok=True)

    processed = 0
    failures = 0
    for order_num, row in rows:
        trial_id = str(cell(ws, row, headers, "trial_id").value)
        prompt = str(cell(ws, row, headers, "prompt_text").value)
        print(f"[{processed + 1}/{len(rows)}] random_order={order_num} trial_id={trial_id}")

        include_sampling = bool(args.include_sampling_params)
        sampling_note = ""
        try:
            try:
                response = make_response(
                    client=client,
                    model=args.model,
                    prompt=prompt,
                    max_output_tokens=args.max_output_tokens,
                    include_sampling=include_sampling,
                    temperature=args.temperature,
                    top_p=args.top_p,
                )
            except Exception as e:
                # Many reasoning-style models do not accept temperature/top_p. Retry once without them.
                if include_sampling:
                    sampling_note = f"sampling params unsupported or failed; retried without temperature/top_p; original error: {type(e).__name__}: {e}"
                    response = make_response(
                        client=client,
                        model=args.model,
                        prompt=prompt,
                        max_output_tokens=args.max_output_tokens,
                        include_sampling=False,
                        temperature=args.temperature,
                        top_p=args.top_p,
                    )
                    include_sampling = False
                else:
                    raise

            response_text = extract_text(response)
            if not response_text:
                response_text = "[EMPTY_RESPONSE_TEXT_EXTRACTED_CHECK_RAW_JSON]"

            cell(ws, row, headers, "execution_date").value = now_iso()
            cell(ws, row, headers, "model_id").value = args.model
            if include_sampling:
                cell(ws, row, headers, "temperature").value = args.temperature
                cell(ws, row, headers, "top_p").value = args.top_p
            else:
                cell(ws, row, headers, "temperature").value = "not_applicable_model_default"
                cell(ws, row, headers, "top_p").value = "not_applicable_model_default"
            cell(ws, row, headers, "system_prompt_used").value = "no"
            cell(ws, row, headers, "conversation_history_used").value = "no"
            cell(ws, row, headers, "tools_used").value = "no"
            cell(ws, row, headers, "browsing_used").value = "no"
            cell(ws, row, headers, "memory_used").value = "no"
            cell(ws, row, headers, "response_text").value = response_text
            cell(ws, row, headers, "response_length_chars").value = len(response_text)

            raw_saved = "no"
            if args.save_raw_json:
                raw_path = raw_dir / f"{int(order_num):04d}_{safe_filename(trial_id)}.json"
                data = response.model_dump() if hasattr(response, "model_dump") else response
                raw_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
                raw_saved = str(raw_path)
            cell(ws, row, headers, "raw_output_saved").value = raw_saved

            old_notes = cell(ws, row, headers, "notes").value
            notes = []
            if old_notes:
                notes.append(str(old_notes))
            if sampling_note:
                notes.append(sampling_note)
            cell(ws, row, headers, "notes").value = " | ".join(notes)

            # Checkpoint after every successful trial.
            wb.save(output_path)
            processed += 1
            if args.sleep_seconds > 0:
                time.sleep(args.sleep_seconds)

        except KeyboardInterrupt:
            print("Interrupted by user. Saving workbook before exit...")
            wb.save(output_path)
            return 130
        except Exception as e:
            failures += 1
            err_msg = f"ERROR during trial {trial_id}: {type(e).__name__}: {e}"
            print(err_msg, file=sys.stderr)
            old_notes = cell(ws, row, headers, "notes").value
            cell(ws, row, headers, "notes").value = (str(old_notes) + " | " if old_notes else "") + err_msg
            wb.save(output_path)
            # Continue to next trial so a single failure does not kill the batch.
            continue

    wb.save(output_path)
    print(f"Done. processed={processed}, failures={failures}, output={output_path}")
    return 0 if failures == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
